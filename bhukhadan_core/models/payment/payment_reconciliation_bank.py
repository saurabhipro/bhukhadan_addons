# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError


class PaymentReconciliationBank(models.Model):
    _name = 'bhu.payment.reconciliation.bank'
    _description = 'Payment Reconciliation (Bank File) / भुगतान समाधान (बैंक फ़ाइल)'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'reconciliation_date desc, name'

    name = fields.Char(string='Reconciliation Number / समाधान संख्या', required=True, default='New', tracking=True)
    reconciliation_date = fields.Date(string='Reconciliation Date / समाधान दिनांक', required=True, 
                                     default=fields.Date.today, tracking=True)
    
    # Related Records
    payment_file_id = fields.Many2one('bhu.payment.file', string='Payment File / भुगतान फ़ाइल', required=True, tracking=True)
    village_id = fields.Many2one('bhu.village', string='Village / ग्राम', related='payment_file_id.village_id', store=True, readonly=True)
    project_id = fields.Many2one('bhu.project', string='Project / परियोजना', related='payment_file_id.project_id', store=True, readonly=True)
    department_id = fields.Many2one('bhu.department', string='Department / विभाग', related='project_id.department_id', store=True, readonly=True)
    
    # Bank File Upload
    bank_file = fields.Binary(string='Bank File / बैंक फ़ाइल', required=True)
    bank_file_filename = fields.Char(string='Bank File Name / बैंक फ़ाइल नाम')
    upload_date = fields.Date(string='Upload Date / अपलोड दिनांक', default=fields.Date.today, tracking=True)
    
    # Reconciliation Lines
    reconciliation_line_ids = fields.One2many('bhu.payment.reconciliation.bank.line', 'reconciliation_id', 
                                             string='Reconciliation Lines / समाधान पंक्तियां')
    
    # Summary
    total_payments = fields.Integer(string='Total Payments / कुल भुगतान', compute='_compute_summary', store=True)
    settled_count = fields.Integer(string='Settled / निपटाया गया', compute='_compute_summary', store=True)
    failed_count = fields.Integer(string='Failed / असफल', compute='_compute_summary', store=True)
    pending_count = fields.Integer(string='Pending / लंबित', compute='_compute_summary', store=True)
    total_amount = fields.Float(string='Total Amount / कुल राशि', compute='_compute_summary', store=True, digits=(16, 2))
    settled_amount = fields.Float(string='Settled Amount / निपटाई गई राशि', compute='_compute_summary', store=True, digits=(16, 2))
    failed_amount = fields.Float(string='Failed Amount / असफल राशि', compute='_compute_summary', store=True, digits=(16, 2))
    
    state = fields.Selection([
        ('draft', 'Draft / प्रारूप'),
        ('processed', 'Processed / प्रसंस्कृत'),
        ('completed', 'Completed / पूर्ण'),
    ], string='Status / स्थिति', default='draft', tracking=True)
    
    @api.depends('reconciliation_line_ids', 'reconciliation_line_ids.status', 'reconciliation_line_ids.credit_amount')
    def _compute_summary(self):
        """Compute reconciliation summary"""
        for record in self:
            lines = record.reconciliation_line_ids
            record.total_payments = len(lines)
            record.settled_count = len(lines.filtered(lambda l: l.status == 'settled'))
            record.failed_count = len(lines.filtered(lambda l: l.status == 'failed'))
            record.pending_count = len(lines.filtered(lambda l: l.status == 'pending'))
            record.total_amount = sum(lines.mapped('credit_amount'))
            record.settled_amount = sum(lines.filtered(lambda l: l.status == 'settled').mapped('credit_amount'))
            record.failed_amount = sum(lines.filtered(lambda l: l.status == 'failed').mapped('credit_amount'))
    
    @api.model_create_multi
    def create(self, vals_list):
        """Generate reconciliation number if not provided"""
        for vals in vals_list:
            if vals.get('name', 'New') == 'New' or not vals.get('name'):
                # Try to use sequence settings from settings master
                # Get project_id from payment_file_id if available
                project_id = vals.get('project_id')
                if not project_id and vals.get('payment_file_id'):
                    payment_file = self.env['bhu.payment.file'].browse(vals['payment_file_id'])
                    if payment_file.exists():
                        project_id = payment_file.project_id.id
                if project_id:
                    sequence_number = self.env['bhuarjan.settings.master'].get_sequence_number(
                        'payment_reconciliation', project_id
                    )
                    if sequence_number:
                        vals['name'] = sequence_number
                    else:
                        # Fallback to ir.sequence
                        vals['name'] = self.env['ir.sequence'].next_by_code('bhu.payment.reconciliation.bank') or 'New'
                else:
                    # No project_id, use fallback
                    vals['name'] = self.env['ir.sequence'].next_by_code('bhu.payment.reconciliation.bank') or 'New'
        return super().create(vals_list)
    
    def action_process_bank_file(self):
        """Process uploaded bank file and match with payment lines"""
        self.ensure_one()
        if not self.bank_file:
            raise ValidationError(_('Please upload bank file first.'))
        
        # Decode bank file
        import base64
        import csv
        import io
        
        try:
            file_content = base64.b64decode(self.bank_file)
            
            # Clear existing lines
            self.reconciliation_line_ids = [(5, 0, 0)]
            line_vals = []

            # Try to determine if it's Excel or CSV
            filename = (self.bank_file_filename or '').lower()
            is_xlsx = filename.endswith('.xlsx')
            is_xls = filename.endswith('.xls')
            is_excel = is_xls or is_xlsx
            
            if is_excel:
                if is_xlsx:
                    # Process XLSX via openpyxl (xlrd no longer supports xlsx).
                    try:
                        from openpyxl import load_workbook
                    except ImportError:
                        raise ValidationError(
                            _("Python library 'openpyxl' is required to process .xlsx files.")
                        )

                    workbook = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
                    sheet = workbook.active
                    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
                    headers = [str(h).strip() if h is not None else '' for h in header_row]

                    for values in sheet.iter_rows(min_row=2, values_only=True):
                        row_data = {}
                        has_data = False
                        for idx, header in enumerate(headers):
                            if not header:
                                continue
                            cell_val = values[idx] if idx < len(values) else ''
                            if cell_val is None:
                                cell_val = ''
                            if cell_val != '':
                                has_data = True
                            row_data[header] = cell_val
                        if has_data:
                            line_vals.append((0, 0, self._prepare_line_vals(row_data)))
                else:
                    # Process XLS via xlrd
                    import xlrd
                    workbook = xlrd.open_workbook(file_contents=file_content)
                    sheet = workbook.sheet_by_index(0)
                    headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
                    
                    for row_idx in range(1, sheet.nrows):
                        row_data = {}
                        has_data = False
                        for col_idx, header in enumerate(headers):
                            if not header:
                                continue
                            value = sheet.cell_value(row_idx, col_idx)
                            if value not in ('', None):
                                has_data = True
                            row_data[header] = value
                        if has_data:
                            line_vals.append((0, 0, self._prepare_line_vals(row_data)))
            else:
                # Process CSV
                if isinstance(file_content, bytes):
                    file_content = file_content.decode('utf-8-sig') # Handle BOM
                
                reader = csv.DictReader(io.StringIO(file_content))
                for row in reader:
                    line_vals.append((0, 0, self._prepare_line_vals(row)))
            
            if line_vals:
                self.reconciliation_line_ids = line_vals
                self.state = 'processed'
                
                # Match with payment file lines
                self._match_payments()
            else:
                raise ValidationError(_('No valid rows found in uploaded bank file.'))
        except Exception as e:
            raise ValidationError(_('Error processing bank file: %s') % str(e))

        lines = self.reconciliation_line_ids
        total_processed = len(lines)
        passed_count = len(lines.filtered(lambda l: l.status == 'settled'))
        failed_count = len(lines.filtered(lambda l: l.status == 'failed'))
        unmatched_count = len(lines.filtered(lambda l: not l.payment_line_id))

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Reconciliation Summary'),
                'message': _(
                    'Processed: %(processed)s | Passed: %(passed)s | Failed: %(failed)s | Unmatched: %(unmatched)s'
                ) % {
                    'processed': total_processed,
                    'passed': passed_count,
                    'failed': failed_count,
                    'unmatched': unmatched_count,
                },
                'type': 'success',
                'sticky': True,
            }
        }

    def _prepare_line_vals(self, row):
        """Helper to map row data to line vals"""
        # Mapping Template 2 columns
        return {
            'utr_number': row.get('UTR Number', '') or row.get('UTR', ''),
            'transaction_reference': (
                row.get('External Ref No', '')
                or row.get('External Ref', '')
                or row.get('Transaction Reference', '')
                or row.get('Transaction Ref', '')
            ),
            'beneficiary_account': str(row.get('Beneficiary Account', '') or row.get('Account Number', '')).split('.')[0],
            'beneficiary_name': row.get('Beneficiary Name', ''),
            'beneficiary_bank_code': row.get('Beneficiary Bank Code', '') or row.get('IFSC Code', ''),
            'credit_amount': float(row.get('Credit Amount', 0) or row.get('Amount', 0) or 0),
            'status': row.get('Status', '').lower() if row.get('Status') else 'pending',
            'event_status': row.get('Event Status', ''),
            'error': row.get('Error', ''),
            'payment_id': row.get('Payment Id', ''),
            'transaction_date': row.get('Date', ''),
        }
    
    def _match_payments(self):
        """Match bank file transactions with payment file lines"""
        self.ensure_one()
        if not self.payment_file_id:
            return
        
        # Get payment file lines
        payment_lines = self.payment_file_id.payment_line_ids
        
        # Match reconciliation lines with payment lines
        for recon_line in self.reconciliation_line_ids:
            matched_payment = self.env['bhu.payment.file.line']
            tx_ref = (recon_line.transaction_reference or recon_line.payment_id or '').strip()
            if tx_ref:
                ExportLine = self.env['bhu.payment.voucher.export.line']
                matched_payment = payment_lines.filtered(
                    lambda p: ExportLine._bank_ref_matches(
                        p.transaction_uuid,
                        p.voucher_export_line_id.external_ref if p.voucher_export_line_id else '',
                        tx_ref,
                    )
                )[:1]
            if not matched_payment:
                matched_payment = payment_lines.filtered(
                    lambda p: p.account_number == recon_line.beneficiary_account and
                    abs(p.net_payable_amount - recon_line.credit_amount) < 0.01
                )[:1]
            
            if matched_payment:
                recon_line.payment_line_id = matched_payment[0].id
                # Determine status based on bank file status
                bank_status = (recon_line.status or '').lower()
                if bank_status == 'executed' or bank_status == 'settled':
                    recon_line.status = 'settled'
                elif recon_line.error or bank_status == 'failed':
                    recon_line.status = 'failed'
                else:
                    recon_line.status = 'pending'
            else:
                recon_line.status = 'pending'

            self.env['bhu.payment.voucher.export.line'].sync_from_reconciliation_line(
                recon_line,
                self.project_id.id,
                self.village_id.id,
            )

    def action_complete_reconciliation(self):
        """Complete reconciliation and update landowner status"""
        self.ensure_one()
        if self.state != 'processed':
            raise ValidationError(_('Please process the file first.'))
        
        for recon_line in self.reconciliation_line_ids:
            if recon_line.payment_line_id and recon_line.payment_line_id.landowner_id:
                landowner = recon_line.payment_line_id.landowner_id
                
                # Find all surveys for this landowner in this village/project
                # We need to update status for each survey since the user wants khasra-wise status
                surveys = self.env['bhu.survey'].search([
                    ('landowner_ids', 'in', landowner.id),
                    ('village_id', '=', self.village_id.id),
                    ('project_id', '=', self.project_id.id)
                ])
                
                for survey in surveys:
                    # Update or create landowner payment status
                    status_val = 'pending'
                    if recon_line.status == 'settled':
                        status_val = 'paid'
                    elif recon_line.status == 'failed':
                        status_val = 'failed'
                    
                    status_record = self.env['bhu.landowner.payment.status'].search([
                        ('landowner_id', '=', landowner.id),
                        ('survey_id', '=', survey.id),
                        ('project_id', '=', self.project_id.id)
                    ], limit=1)
                    
                    vals = {
                        'landowner_id': landowner.id,
                        'survey_id': survey.id,
                        'project_id': self.project_id.id,
                        'village_id': self.village_id.id,
                        'payment_file_id': self.payment_file_id.id,
                        'utr_number': recon_line.utr_number,
                        'transaction_date': fields.Date.today(), # Or extract from file
                        'amount': recon_line.credit_amount,
                        'status': status_val,
                        'remarks': recon_line.error or recon_line.event_status or ''
                    }
                    
                    if status_record:
                        status_record.write(vals)
                    else:
                        self.env['bhu.landowner.payment.status'].create(vals)
        
        self.state = 'completed'
        return True


class PaymentReconciliationBankLine(models.Model):
    _name = 'bhu.payment.reconciliation.bank.line'
    _description = 'Payment Reconciliation Bank Line / भुगतान समाधान बैंक पंक्ति'
    _order = 'utr_number'

    reconciliation_id = fields.Many2one('bhu.payment.reconciliation.bank', string='Reconciliation', 
                                        required=True, ondelete='cascade')
    
    # Bank File Data
    utr_number = fields.Char(string='UTR Number / यूटीआर संख्या')
    transaction_reference = fields.Char(string='Transaction Reference / लेनदेन संदर्भ')
    beneficiary_account = fields.Char(string='Beneficiary Account / लाभार्थी खाता')
    beneficiary_name = fields.Char(string='Beneficiary Name / लाभार्थी नाम')
    beneficiary_bank_code = fields.Char(string='Beneficiary Bank Code / लाभार्थी बैंक कोड')
    credit_amount = fields.Float(string='Credit Amount / क्रेडिट राशि', digits=(16, 2))
    status = fields.Selection([
        ('pending', 'Pending / लंबित'),
        ('settled', 'Settled / निपटाया गया'),
        ('failed', 'Failed / असफल'),
    ], string='Status / स्थिति', default='pending')
    event_status = fields.Char(string='Event Status / इवेंट स्थिति')
    error = fields.Text(string='Error / त्रुटि')
    payment_id = fields.Char(string='Payment ID / भुगतान आईडी')
    transaction_date = fields.Char(string='Transaction Date / लेनदेन दिनांक')
    
    # Matched Payment Line
    payment_line_id = fields.Many2one('bhu.payment.file.line', string='Matched Payment Line / मिलान भुगतान पंक्ति')
    
    # Computed fields from matched payment
    expected_amount = fields.Float(string='Expected Amount / अपेक्षित राशि', 
                                  related='payment_line_id.net_payable_amount', readonly=True, digits=(16, 2))
    amount_difference = fields.Float(string='Amount Difference / राशि अंतर', 
                                    compute='_compute_amount_difference', store=True, digits=(16, 2))
    
    @api.depends('credit_amount', 'expected_amount')
    def _compute_amount_difference(self):
        """Compute difference between expected and actual amount"""
        for record in self:
            if record.expected_amount and record.credit_amount:
                record.amount_difference = record.credit_amount - record.expected_amount
            else:
                record.amount_difference = 0.0


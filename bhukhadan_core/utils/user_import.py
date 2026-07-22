# -*- coding: utf-8 -*-
"""Import Tehsildar, SDM, and Patwari users from the government roster XLSX layout.

Resolves Project / Department / Tehsil / Sub Division / Village masters. When a
master name is missing and a district is selected, creates the master then maps
the user onto it.
"""

import base64
import io
import re

from odoo.exceptions import ValidationError

# 0-based column indexes when headers match the Patwari export / Google Sheet layout.
_DEFAULT_COL_MAP = {
    'project': None,
    'department': None,
    'department_user': None,
    'tehsil': 1,
    'tehsildar': 2,
    'sub_division': 3,
    'sdm': 4,
    'village': 7,
    'patwari': 8,
    'mobile': 9,
    'email': 10,
}

_ROLE_LABELS = {
    'tehsildar': 'Tehsildar',
    'nodal_officer_lr': 'SDM',
    'halka_patwari': 'Patwari',
    'staff_officer_pp': 'Department User',
}


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _normalize_mobile(val):
    digits = ''.join(c for c in str(val or '') if c.isdigit())
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def _normalize_name_key(name):
    return re.sub(r'\s+', ' ', (name or '').strip()).casefold()


def _strip_bilingual_label(text):
    raw = _cell_text(text)
    if not raw:
        return ''
    if '/' in raw:
        parts = [p.strip() for p in raw.split('/') if p.strip()]
        if parts:
            return parts[0]
    return raw


def _extract_bracket_code(cell_text):
    if not cell_text:
        return ''
    match = re.match(r'^\s*\[([^\]]+)\]', cell_text.strip())
    return match.group(1).strip() if match else ''


def _plain_label(cell_text):
    """Name part after optional ``[CODE]`` prefix (also strips bilingual ``/``)."""
    if not cell_text:
        return ''
    plain = re.sub(r'^\[[^\]]+\]\s*', '', cell_text.strip()).strip()
    return _strip_bilingual_label(plain)


def _plain_village_label(cell_text):
    return _plain_label(cell_text)


def _parse_coded_cell(cell_text):
    """Return ``(code, plain_name)`` from values like ``[V1] Village``."""
    raw = _cell_text(cell_text)
    code = _extract_bracket_code(raw)
    plain = _plain_label(raw)
    return code, plain


def _next_master_code(env, model, field_name, prefix):
    """Next free code like V1, P2, D3 for the given model/field."""
    Model = env[model].sudo()
    prefix = (prefix or '').upper()
    pattern = re.compile(rf'^{re.escape(prefix)}(\d+)$', re.IGNORECASE)
    max_n = 0
    # Limit scan — codes are short; fetch existing non-empty codes with prefix
    records = Model.search([(field_name, '=ilike', f'{prefix}%')])
    for rec in records:
        val = (rec[field_name] or '').strip()
        match = pattern.match(val)
        if match:
            max_n = max(max_n, int(match.group(1)))
    candidate_n = max_n + 1
    while Model.search_count([(field_name, '=ilike', f'{prefix}{candidate_n}')]):
        candidate_n += 1
    return f'{prefix}{candidate_n}'


def _resolve_code(env, model, field_name, prefix, excel_code, existing=None):
    """Prefer Excel ``[CODE]``, else keep existing, else auto-generate."""
    excel_code = (excel_code or '').strip()
    if excel_code:
        return excel_code
    if existing and existing[field_name]:
        return (existing[field_name] or '').strip()
    return _next_master_code(env, model, field_name, prefix)


def _backfill_code(record, field_name, code, dry_run, log_lines, label):
    """Set code on an existing master when it was blank."""
    if not record or not code or not field_name:
        return
    current = (record[field_name] or '').strip()
    if current:
        return
    if dry_run:
        if log_lines is not None:
            log_lines.append(f'  Would set {label} code [{code}] on "{record.display_name}"')
        return
    record.sudo().write({field_name: code})
    if log_lines is not None:
        log_lines.append(f'  Set {label} code [{code}] on "{record.display_name}"')


def _detect_column_map(headers):
    col_map = dict(_DEFAULT_COL_MAP)
    for idx, raw in enumerate(headers):
        cell = raw or ''
        low = cell.strip().lower()
        if 'project' in low or 'परियोजना' in cell or 'प्रोजेक्ट' in cell:
            col_map['project'] = idx
        elif (
            ('department' in low and 'user' not in low)
            or cell.strip() in ('विभाग', 'Department')
            or ('विभाग' in cell and 'उपयोग' not in cell and 'user' not in low)
        ):
            col_map['department'] = idx
        elif (
            'department user' in low
            or 'dept user' in low
            or 'विभाग उपयोग' in cell
            or low in ('department officer', 'dept officer')
        ):
            col_map['department_user'] = idx
        elif 'तहसील' in cell and 'तहसीलदार' not in cell:
            col_map['tehsil'] = idx
        elif 'तहसीलदार' in cell or low == 'tehsildar':
            col_map['tehsildar'] = idx
        elif 'sub division' in low or 'उपभाग' in cell or 'अनुविभाग' in cell:
            col_map['sub_division'] = idx
        elif low == 'sdm' or cell.strip() == 'SDM':
            col_map['sdm'] = idx
        elif 'ग्राम' in cell or 'village' in low or 'प्रभावित' in cell:
            col_map['village'] = idx
        elif 'पटवारी' in cell and 'मोबाइल' not in cell:
            col_map['patwari'] = idx
        elif 'mobile' in low or 'मोबाइल' in cell:
            col_map['mobile'] = idx
        elif 'email' in low or 'e-mail' in low:
            col_map['email'] = idx
    return col_map


def _load_xlsx_rows(file_content, filename):
    try:
        from openpyxl import load_workbook
    except ImportError as err:
        raise ValidationError(
            "Python library 'openpyxl' is required to import .xlsx files."
        ) from err

    raw = base64.b64decode(file_content)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_cell_text(c) for c in row])
    wb.close()
    if not rows:
        raise ValidationError('The Excel file is empty.')
    return rows


def _pick_best_name_match(records, plain):
    """Prefer exact (casefold) name, then shortest ilike hit."""
    if not records:
        return records.browse()
    key = _normalize_name_key(plain)
    exact = records.filtered(lambda r: _normalize_name_key(r.name) == key)
    if exact:
        return exact[0]
    return records.sorted(key=lambda r: len(r.name or ''))[:1]


def _find_master_by_name(env, model, label, district_id=None, code=None, code_field='code'):
    plain = _strip_bilingual_label(label) if label else ''
    Model = env[model].sudo()

    if code and code_field in Model._fields:
        domain = [(code_field, '=ilike', code)]
        if district_id and 'district_id' in Model._fields:
            domain.append(('district_id', '=', district_id))
        record = Model.search(domain, limit=1)
        if record:
            return record
        # Code is global-unique for several masters — try without district
        record = Model.search([(code_field, '=ilike', code)], limit=1)
        if record:
            return record

    if not plain:
        return Model.browse()

    domain = [('name', 'ilike', plain)]
    if district_id and 'district_id' in Model._fields:
        scoped = Model.search(domain + [('district_id', '=', district_id)])
        picked = _pick_best_name_match(scoped, plain)
        if picked:
            return picked

    return _pick_best_name_match(Model.search(domain), plain)


def _find_village(env, cell_text, district_id=None, tehsil=None, subdiv=None):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.village'].browse()

    Village = env['bhu.village'].sudo()
    if code:
        domain = [('village_code', '=ilike', code)]
        if district_id:
            domain.append(('district_id', '=', district_id))
        village = Village.search(domain, limit=1)
        if village:
            return village
        village = Village.search([('village_code', '=ilike', code)], limit=1)
        if village:
            return village

    if plain:
        domain = ['|', ('name', '=ilike', plain), ('name', 'ilike', plain)]
        extras = []
        if district_id:
            extras.append(('district_id', '=', district_id))
        if tehsil:
            extras.append(('tehsil_id', '=', tehsil.id))
        if subdiv:
            extras.append(('sub_division_id', '=', subdiv.id))
        if extras:
            scoped = Village.search(extras + domain)
            picked = _pick_best_name_match(scoped, plain)
            if picked:
                return picked
        return _pick_best_name_match(Village.search(domain), plain)
    return env['bhu.village'].browse()


def _find_project(env, cell_text, district_id=None):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.project'].browse()

    Project = env['bhu.project'].sudo()
    if code:
        domain = [('code', '=ilike', code)]
        if district_id:
            domain.append(('district_id', '=', district_id))
        project = Project.search(domain, limit=1)
        if project:
            return project
        project = Project.search([('code', '=ilike', code)], limit=1)
        if project:
            return project

    domain = ['|', ('name', '=ilike', plain), ('name', 'ilike', plain)]
    if district_id:
        scoped = Project.search([('district_id', '=', district_id)] + domain)
        picked = _pick_best_name_match(scoped, plain)
        if picked:
            return picked
    return _pick_best_name_match(Project.search(domain), plain)


def _ensure_sub_division(env, cell_text, district_id, state_id, create_missing, dry_run, log_lines):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.sub.division'].browse(), False
    record = _find_master_by_name(
        env, 'bhu.sub.division', plain, district_id, code=code, code_field='code',
    )
    if record:
        resolved = _resolve_code(env, 'bhu.sub.division', 'code', 'SD', code, existing=record)
        _backfill_code(record, 'code', resolved, dry_run, log_lines, 'Sub Division')
        return record, False
    if not create_missing:
        if log_lines is not None:
            log_lines.append(f'  Warning: Sub Division not found: {cell_text}')
        return env['bhu.sub.division'].browse(), False
    if not district_id:
        if log_lines is not None:
            log_lines.append(
                f'  Warning: Cannot create Sub Division "{plain or code}" — select a District on the wizard.'
            )
        return env['bhu.sub.division'].browse(), False
    district = env['bhu.district'].browse(district_id)
    resolved = _resolve_code(env, 'bhu.sub.division', 'code', 'SD', code)
    vals = {
        'name': plain or resolved,
        'code': resolved,
        'district_id': district_id,
        'state_id': state_id or district.state_id.id,
    }
    if dry_run:
        if log_lines is not None:
            log_lines.append(f'  Would create Sub Division: [{resolved}] {vals["name"]}')
        return env['bhu.sub.division'].browse(), True
    record = env['bhu.sub.division'].sudo().create(vals)
    if log_lines is not None:
        log_lines.append(f'  Created Sub Division: {record.display_name}')
    return record, True


def _ensure_tehsil(env, cell_text, district_id, state_id, subdiv, create_missing, dry_run, log_lines):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.tehsil'].browse(), False
    record = _find_master_by_name(
        env, 'bhu.tehsil', plain, district_id, code=code, code_field='code',
    )
    if record:
        resolved = _resolve_code(env, 'bhu.tehsil', 'code', 'T', code, existing=record)
        _backfill_code(record, 'code', resolved, dry_run, log_lines, 'Tehsil')
        return record, False
    if not create_missing:
        if log_lines is not None:
            log_lines.append(f'  Warning: Tehsil not found: {cell_text}')
        return env['bhu.tehsil'].browse(), False
    if not district_id:
        if log_lines is not None:
            log_lines.append(
                f'  Warning: Cannot create Tehsil "{plain or code}" — select a District on the wizard.'
            )
        return env['bhu.tehsil'].browse(), False
    district = env['bhu.district'].browse(district_id)
    resolved = _resolve_code(env, 'bhu.tehsil', 'code', 'T', code)
    vals = {
        'name': plain or resolved,
        'code': resolved,
        'district_id': district_id,
        'state_id': state_id or district.state_id.id,
    }
    if subdiv:
        vals['sub_division_id'] = subdiv.id
    if dry_run:
        if log_lines is not None:
            log_lines.append(f'  Would create Tehsil: [{resolved}] {vals["name"]}')
        return env['bhu.tehsil'].browse(), True
    record = env['bhu.tehsil'].sudo().create(vals)
    if log_lines is not None:
        log_lines.append(f'  Created Tehsil: {record.display_name}')
    return record, True


def _ensure_village(env, cell_text, district_id, state_id, tehsil, subdiv,
                    create_missing, dry_run, log_lines):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.village'].browse(), False

    record = _find_village(env, cell_text, district_id, tehsil=tehsil, subdiv=subdiv)
    if record:
        resolved = _resolve_code(env, 'bhu.village', 'village_code', 'V', code, existing=record)
        _backfill_code(record, 'village_code', resolved, dry_run, log_lines, 'Village')
        return record, False
    if not create_missing:
        if log_lines is not None:
            log_lines.append(f'  Warning: Village not found: {cell_text}')
        return env['bhu.village'].browse(), False
    if not district_id:
        if log_lines is not None:
            log_lines.append(
                f'  Warning: Cannot create Village "{plain or cell_text}" — '
                f'select a District on the wizard.'
            )
        return env['bhu.village'].browse(), False

    district = env['bhu.district'].browse(district_id)
    resolved = _resolve_code(env, 'bhu.village', 'village_code', 'V', code)
    vals = {
        'name': plain or resolved,
        'village_code': resolved,
        'district_id': district_id,
        'state_id': state_id or district.state_id.id,
    }
    if tehsil:
        vals['tehsil_id'] = tehsil.id
    if subdiv:
        vals['sub_division_id'] = subdiv.id
    elif tehsil and tehsil.sub_division_id:
        vals['sub_division_id'] = tehsil.sub_division_id.id

    if dry_run:
        if log_lines is not None:
            log_lines.append(f'  Would create Village: [{resolved}] {vals["name"]}')
        return env['bhu.village'].browse(), True
    record = env['bhu.village'].sudo().create(vals)
    if log_lines is not None:
        log_lines.append(f'  Created Village: {record.display_name}')
    return record, True


def _find_department(env, cell_text, district_id=None):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.department'].browse()

    Department = env['bhu.department'].sudo()
    if code:
        domain = [('code', '=ilike', code)]
        if district_id:
            domain.append(('district_id', '=', district_id))
        dept = Department.search(domain, limit=1)
        if dept:
            return dept
        dept = Department.search([('code', '=ilike', code)], limit=1)
        if dept:
            return dept

    domain = ['|', ('name', '=ilike', plain), ('name', 'ilike', plain)]
    if district_id:
        scoped = Department.search([('district_id', '=', district_id)] + domain)
        picked = _pick_best_name_match(scoped, plain)
        if picked:
            return picked
    return _pick_best_name_match(Department.search(domain), plain)


def _ensure_department(env, cell_text, district_id, create_missing, dry_run, log_lines):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.department'].browse(), False

    record = _find_department(env, cell_text, district_id)
    if record:
        resolved = _resolve_code(env, 'bhu.department', 'code', 'D', code, existing=record)
        _backfill_code(record, 'code', resolved, dry_run, log_lines, 'Department')
        return record, False
    if not create_missing:
        if log_lines is not None:
            log_lines.append(f'  Warning: Department not found: {cell_text}')
        return env['bhu.department'].browse(), False

    resolved = _resolve_code(env, 'bhu.department', 'code', 'D', code)
    vals = {
        'name': plain or resolved,
        'code': resolved,
        'active': True,
    }
    if district_id:
        vals['district_id'] = district_id

    if dry_run:
        if log_lines is not None:
            log_lines.append(f'  Would create Department: [{resolved}] {vals["name"]}')
        return env['bhu.department'].browse(), True
    record = env['bhu.department'].sudo().create(vals)
    if log_lines is not None:
        log_lines.append(f'  Created Department: {record.display_name}')
    return record, True


def _ensure_project(env, cell_text, district_id, department, create_missing, dry_run, log_lines):
    code, plain = _parse_coded_cell(cell_text)
    if not plain and not code:
        return env['bhu.project'].browse(), False

    record = _find_project(env, cell_text, district_id)
    if record:
        resolved = _resolve_code(env, 'bhu.project', 'code', 'P', code, existing=record)
        _backfill_code(record, 'code', resolved, dry_run, log_lines, 'Project')
        # Attach department on existing project when empty
        if department and not record.department_id and not dry_run:
            record.sudo().write({'department_id': department.id})
            if log_lines is not None:
                log_lines.append(
                    f'  Set Project "{record.display_name}" department → {department.display_name}'
                )
        return record, False
    if not create_missing:
        if log_lines is not None:
            log_lines.append(f'  Warning: Project not found: {cell_text}')
        return env['bhu.project'].browse(), False

    company = env.user.company_id or env.company
    resolved = _resolve_code(env, 'bhu.project', 'code', 'P', code)
    vals = {
        'name': plain or resolved,
        'code': resolved,
        'state': 'active',
    }
    if district_id:
        vals['district_id'] = district_id
    if department:
        vals['department_id'] = department.id
    if company:
        vals['company_id'] = company.id

    if dry_run:
        if log_lines is not None:
            log_lines.append(f'  Would create Project: [{resolved}] {vals["name"]}')
        return env['bhu.project'].browse(), True
    record = env['bhu.project'].sudo().create(vals)
    if log_lines is not None:
        log_lines.append(f'  Created Project: {record.display_name}')
    return record, True


def _link_village_to_project(project, village, dry_run, log_lines):
    if not project or not village or not project.id or not village.id:
        return False
    if village in project.village_ids:
        return False
    if dry_run:
        if log_lines is not None:
            log_lines.append(
                f'  Would map Village "{village.display_name}" → Project "{project.display_name}"'
            )
        return True
    project.sudo().write({'village_ids': [(4, village.id)]})
    if log_lines is not None:
        log_lines.append(
            f'  Mapped Village "{village.display_name}" → Project "{project.display_name}"'
        )
    return True


def _link_user_to_project_department(project, user, dry_run, log_lines):
    """Add user on project.department_user_ids when that field exists."""
    if not project or not user or not project.id or not user.id:
        return False
    if 'department_user_ids' not in project._fields:
        return False
    if user in project.department_user_ids:
        return False
    if dry_run:
        if log_lines is not None:
            log_lines.append(
                f'  Would add Department User "{user.name}" on Project "{project.display_name}"'
            )
        return True
    project.sudo().write({'department_user_ids': [(4, user.id)]})
    if log_lines is not None:
        log_lines.append(
            f'  Added Department User "{user.name}" on Project "{project.display_name}"'
        )
    return True


def _assign_bhuarjan_groups(env, user, role):
    Users = env['res.users']
    role_groups = Users._bhuarjan_role_group_xml_ids()
    all_custom = Users._bhuarjan_all_group_xml_ids()
    all_group_ids = []
    for xid in all_custom:
        group = env.ref(xid, raise_if_not_found=False)
        if group:
            all_group_ids.append(group.id)

    current_ids = user.groups_id.ids
    kept = [gid for gid in current_ids if gid not in all_group_ids]
    group_ref = role_groups.get(role)
    if group_ref:
        group = env.ref(group_ref, raise_if_not_found=False)
        if group and group.id not in kept:
            kept.append(group.id)
    base_user = env.ref('base.group_user', raise_if_not_found=False)
    if base_user and base_user.id not in kept:
        kept.append(base_user.id)
    user.sudo().write({'groups_id': [(6, 0, kept)]})


def _make_login(env, role, name, email=None, mobile=None, district_id=None):
    if email:
        login = email.strip().lower()
        if login:
            return login

    slug = re.sub(r'[^a-z0-9]+', '.', _normalize_name_key(name)).strip('.') or 'user'
    prefix = {
        'halka_patwari': 'patwari',
        'tehsildar': 'tehsildar',
        'nodal_officer_lr': 'sdm',
        'staff_officer_pp': 'dept',
    }.get(role, 'user')
    if mobile:
        base = f'{prefix}.{mobile}'
    else:
        base = f'{prefix}.{slug}'
    if district_id:
        base = f'{base}.d{district_id}'
    candidate = f'{base}@import.bhuarjan'
    suffix = 1
    Users = env['res.users'].sudo()
    while Users.search_count([('login', '=', candidate)]):
        suffix += 1
        candidate = f'{base}{suffix}@import.bhuarjan'
    return candidate


def _find_existing_user(env, role, name, login=None, mobile=None, district_id=None):
    Users = env['res.users'].sudo()
    if mobile:
        user = Users.search([('mobile', '=', mobile)], limit=1)
        if user:
            return user
    if login:
        user = Users.search([('login', '=', login)], limit=1)
        if user:
            return user
    name_key = _normalize_name_key(name)
    if not name_key:
        return Users.browse()
    domain = [('bhuarjan_role', '=', role), ('name', 'ilike', name.strip())]
    if district_id:
        domain.append(('district_id', '=', district_id))
    candidates = Users.search(domain)
    for user in candidates:
        if _normalize_name_key(user.name) == name_key:
            return user
    return Users.browse()


def _ensure_user(env, cache, role, name, email=None, mobile=None, district_id=None,
                 state_id=None, update_existing=False, dry_run=False, log_lines=None):
    name = (name or '').strip()
    if not name:
        return None, 'skipped', None

    mobile = _normalize_mobile(mobile) or False
    login = (email or '').strip().lower() or None
    cache_key = (role, mobile or login or _normalize_name_key(name))

    if cache_key in cache:
        cached = cache[cache_key]
        return cached, 'cached', cached.name if cached else name

    existing = _find_existing_user(env, role, name, login=login, mobile=mobile, district_id=district_id)
    vals = {'name': name, 'bhuarjan_role': role}
    if mobile:
        vals['mobile'] = mobile
    if district_id:
        vals['district_id'] = district_id
    if state_id:
        vals['state_id'] = state_id

    if existing:
        if update_existing:
            if not dry_run:
                existing.sudo().write(vals)
                _assign_bhuarjan_groups(env, existing, role)
            cache[cache_key] = existing
            if log_lines is not None:
                log_lines.append(f'  Updated {_ROLE_LABELS[role]}: {name}')
            return existing, 'updated', name
        cache[cache_key] = existing
        return existing, 'existing', name

    login = _make_login(env, role, name, email=login, mobile=mobile, district_id=district_id)
    vals['login'] = login
    vals['active'] = True
    company = env.user.company_id or env.company
    if company:
        vals['company_id'] = company.id
        vals['company_ids'] = [(6, 0, [company.id])]

    if dry_run:
        cache[cache_key] = existing or env['res.users'].browse()
        if log_lines is not None:
            log_lines.append(f'  Would create {_ROLE_LABELS[role]}: {name} ({login})')
        return cache[cache_key], 'created', name

    user = env['res.users'].sudo().create(vals)
    _assign_bhuarjan_groups(env, user, role)
    cache[cache_key] = user
    if log_lines is not None:
        log_lines.append(f'  Created {_ROLE_LABELS[role]}: {name} ({login})')
    return user, 'created', name


def _link_master(master, user, field_name, dry_run, log_lines, label,
                 user_display_name=None, force_relink=False):
    if not master:
        return False
    display_name = user_display_name or (user.name if user else '')
    if not display_name:
        return False
    current = master[field_name]
    if current and user and current.id != user.id:
        if not force_relink:
            if log_lines is not None:
                log_lines.append(
                    f'  Warning: {label} "{master.display_name}" already linked to '
                    f'"{current.display_name}" — skipped relink to "{display_name}".'
                )
            return False
        if log_lines is not None:
            log_lines.append(
                f'  Relinking {label} "{master.display_name}" from '
                f'"{current.display_name}" → "{display_name}"'
            )
    if current and user and current.id == user.id:
        return False
    if not dry_run and user and user.id:
        master.sudo().write({field_name: user.id})
    if log_lines is not None:
        action = 'Would link' if dry_run else 'Linked'
        log_lines.append(f'  {action} {label} "{master.display_name}" → {display_name}')
    return True


def import_user_roster_xlsx(env, file_content, filename, district_id=None,
                            update_existing=False, dry_run=False,
                            create_missing_masters=True, force_relink=False):
    """Parse roster XLSX and create/link Tehsildar, SDM, and Patwari users.

    When ``create_missing_masters`` is True and a District is set, missing
    Project / Department / Tehsil / Sub Division / Village rows are created, then users are mapped.
    """
    stats = {
        'created': 0,
        'updated': 0,
        'tehsils_linked': 0,
        'subdivisions_linked': 0,
        'villages_linked': 0,
        'projects_linked': 0,
        'departments_linked': 0,
        'masters_created': 0,
        'rows': 0,
        'errors': 0,
        'skipped': 0,
    }
    log_lines = []
    prefix = '[DRY RUN] ' if dry_run else ''
    log_lines.append(f'{prefix}User roster import started for file: {filename or "upload.xlsx"}')
    log_lines.append(
        f'Options: create_missing_masters={create_missing_masters}, '
        f'force_relink={force_relink}, update_existing={update_existing}'
    )

    rows = _load_xlsx_rows(file_content, filename)
    headers = rows[0]
    col_map = _detect_column_map(headers)
    log_lines.append(
        f'Columns: project={col_map.get("project")}, department={col_map.get("department")}, '
        f'dept_user={col_map.get("department_user")}, tehsil={col_map["tehsil"]}, '
        f'tehsildar={col_map["tehsildar"]}, sub_div={col_map["sub_division"]}, '
        f'sdm={col_map["sdm"]}, village={col_map["village"]}, '
        f'patwari={col_map["patwari"]}, mobile={col_map["mobile"]}, email={col_map["email"]}'
    )

    district = env['bhu.district'].browse(district_id) if district_id else env['bhu.district'].browse()
    state_id = district.state_id.id if district else False

    user_cache = {}

    def cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ''
        return row[idx]

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or not any(c for c in row):
            continue

        project_cell = cell(row, 'project')
        department_cell = cell(row, 'department')
        department_user_name = cell(row, 'department_user')
        tehsil_name = cell(row, 'tehsil')
        tehsildar_name = cell(row, 'tehsildar')
        subdiv_name = cell(row, 'sub_division')
        sdm_name = cell(row, 'sdm')
        village_cell = cell(row, 'village')
        patwari_name = cell(row, 'patwari')
        mobile = cell(row, 'mobile')
        email = cell(row, 'email')

        if not any([
            tehsildar_name, sdm_name, patwari_name, project_cell,
            department_cell, department_user_name,
        ]):
            continue

        stats['rows'] += 1
        log_lines.append(f'Row {row_num}:')

        try:
            department, created = _ensure_department(
                env, department_cell, district_id,
                create_missing_masters, dry_run, log_lines,
            ) if department_cell else (env['bhu.department'].browse(), False)
            if created:
                stats['masters_created'] += 1

            subdiv, created = _ensure_sub_division(
                env, subdiv_name, district_id, state_id,
                create_missing_masters, dry_run, log_lines,
            )
            if created:
                stats['masters_created'] += 1

            tehsil, created = _ensure_tehsil(
                env, tehsil_name, district_id, state_id, subdiv,
                create_missing_masters, dry_run, log_lines,
            )
            if created:
                stats['masters_created'] += 1

            village, created = _ensure_village(
                env, village_cell, district_id, state_id, tehsil, subdiv,
                create_missing_masters, dry_run, log_lines,
            )
            if created:
                stats['masters_created'] += 1

            project, created = _ensure_project(
                env, project_cell, district_id, department,
                create_missing_masters, dry_run, log_lines,
            ) if project_cell else (env['bhu.project'].browse(), False)
            if created:
                stats['masters_created'] += 1

            # If we have department + project but project had no dept set above path for new only
            if project and department and project.department_id.id != department.id:
                if not project.department_id or force_relink:
                    if not dry_run:
                        project.sudo().write({'department_id': department.id})
                    if log_lines is not None:
                        action = 'Would set' if dry_run else 'Set'
                        log_lines.append(
                            f'  {action} Project "{project.display_name}" department → '
                            f'{department.display_name}'
                        )

            row_district_id = district_id
            row_state_id = state_id
            if village and village.district_id:
                row_district_id = village.district_id.id
                row_state_id = village.state_id.id or row_state_id
            elif tehsil and tehsil.district_id:
                row_district_id = tehsil.district_id.id
                row_state_id = tehsil.state_id.id or row_state_id
            elif subdiv and subdiv.district_id:
                row_district_id = subdiv.district_id.id
                row_state_id = subdiv.state_id.id or row_state_id

            if department_user_name:
                user, status, display_name = _ensure_user(
                    env, user_cache, 'staff_officer_pp', department_user_name,
                    email=email if not patwari_name else None,
                    mobile=mobile if not patwari_name else None,
                    district_id=row_district_id, state_id=row_state_id,
                    update_existing=update_existing, dry_run=dry_run, log_lines=log_lines,
                )
                if status == 'created':
                    stats['created'] += 1
                elif status == 'updated':
                    stats['updated'] += 1
                if department and display_name and _link_master(
                    department, user, 'user_id', dry_run, log_lines, 'Department',
                    display_name, force_relink=force_relink,
                ):
                    stats['departments_linked'] += 1
                if project and user and _link_user_to_project_department(
                    project, user, dry_run, log_lines,
                ):
                    stats['departments_linked'] += 1

            if tehsildar_name:
                user, status, display_name = _ensure_user(
                    env, user_cache, 'tehsildar', tehsildar_name,
                    district_id=row_district_id, state_id=row_state_id,
                    update_existing=update_existing, dry_run=dry_run, log_lines=log_lines,
                )
                if status == 'created':
                    stats['created'] += 1
                elif status == 'updated':
                    stats['updated'] += 1
                if tehsil and display_name and _link_master(
                    tehsil, user, 'user_id', dry_run, log_lines, 'Tehsil',
                    display_name, force_relink=force_relink,
                ):
                    stats['tehsils_linked'] += 1

            if sdm_name:
                user, status, display_name = _ensure_user(
                    env, user_cache, 'nodal_officer_lr', sdm_name,
                    district_id=row_district_id, state_id=row_state_id,
                    update_existing=update_existing, dry_run=dry_run, log_lines=log_lines,
                )
                if status == 'created':
                    stats['created'] += 1
                elif status == 'updated':
                    stats['updated'] += 1
                if subdiv and display_name and _link_master(
                    subdiv, user, 'user_id', dry_run, log_lines, 'Sub Division',
                    display_name, force_relink=force_relink,
                ):
                    stats['subdivisions_linked'] += 1

            if patwari_name:
                user, status, display_name = _ensure_user(
                    env, user_cache, 'halka_patwari', patwari_name,
                    email=email, mobile=mobile,
                    district_id=row_district_id, state_id=row_state_id,
                    update_existing=update_existing, dry_run=dry_run, log_lines=log_lines,
                )
                if status == 'created':
                    stats['created'] += 1
                elif status == 'updated':
                    stats['updated'] += 1
                if village and display_name and _link_master(
                    village, user, 'user_id', dry_run, log_lines, 'Village',
                    display_name, force_relink=force_relink,
                ):
                    stats['villages_linked'] += 1

            if project and village and _link_village_to_project(
                project, village, dry_run, log_lines,
            ):
                stats['projects_linked'] += 1

        except Exception as err:
            stats['errors'] += 1
            log_lines.append(f'  ERROR: {err}')

    summary = (
        f'{prefix}Done — rows: {stats["rows"]}, users created: {stats["created"]}, '
        f'updated: {stats["updated"]}, masters created: {stats["masters_created"]}, '
        f'departments linked: {stats["departments_linked"]}, '
        f'tehsils linked: {stats["tehsils_linked"]}, '
        f'sub divisions linked: {stats["subdivisions_linked"]}, '
        f'villages linked: {stats["villages_linked"]}, '
        f'project↔village maps: {stats["projects_linked"]}, errors: {stats["errors"]}.'
    )
    log_lines.insert(1, summary)
    return stats, '\n'.join(log_lines)

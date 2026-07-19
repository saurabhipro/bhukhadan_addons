# -*- coding: utf-8 -*-
"""Parse BhuKhadan acquisition XLSX and import surveys with khata/landowner masters."""

import base64
import io
import re
from datetime import date

from odoo.exceptions import ValidationError

from odoo.addons.bhukhadan_core.utils.survey_api import api_build_survey_vals

# Fallback column layout (0-based) when headers are not detected.
_DEFAULT_COL_MAP = {
    'khata': 2,
    'owner': 3,
    'khasra': 4,
    'land_type': 5,
    'area_ha': 6,
    'acquired_ha': 7,
    'khata_total_ha': 8,
    'khata_total_acre': 9,
    'acq_reference': 10,
    'caste': 11,
    'employment': 12,
}

_HEADER_MARKERS = ('खाता', 'khasra', 'khasra_number', 'khata', 'plot', 'प्ल')


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _cell_float(value, default=0.0):
    text = _cell_text(value).replace(',', '')
    if not text:
        return default
    try:
        return float(text)
    except (TypeError, ValueError):
        return default


def _normalize_header(text):
    return re.sub(r'\s+', ' ', _cell_text(text).lower())


def _score_khasra_header(header):
    score = 0
    if 'खसरा' in header or 'khasra' in header:
        score += 15
    if 'प्ल' in header or 'plot' in header or 'plant' in header:
        score += 12
    if 'भूमिस्वामी' in header or 'landowner' in header or 'पिता' in header or 'पति' in header:
        score -= 30
    return score


def _score_owner_header(header):
    score = 0
    if 'भूमिस्वामी' in header or 'landowner' in header:
        score += 20
    if 'पिता' in header or 'पति' in header:
        score += 10
    if 'खसरा' in header or 'khasra' in header or 'प्ल' in header or 'plot' in header:
        score -= 30
    return score


def _score_simple_header(header, markers, negative=()):
    score = 0
    for marker in markers:
        if marker in header:
            score += 10
    for marker in negative:
        if marker in header:
            score -= 20
    return score


def _detect_column_map(sheet, header_row):
    """Map logical fields to column indexes using Hindi/English header labels."""
    headers = [
        _normalize_header(sheet.cell(header_row, col_idx + 1).value)
        for col_idx in range(sheet.max_column)
    ]

    def _best_column(score_fn):
        best_col = None
        best_score = 0
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            score = score_fn(header)
            if score > best_score:
                best_score = score
                best_col = col_idx
        return best_col

    col_map = {
        'khata': _best_column(lambda h: _score_simple_header(h, ('खाता', 'khata'))),
        'owner': _best_column(_score_owner_header),
        'khasra': _best_column(_score_khasra_header),
        'land_type': _best_column(lambda h: _score_simple_header(
            h, ('भूमि', 'land type', 'प्रकार'), ('खसरा', 'khasra'),
        )),
        'area_ha': _best_column(lambda h: _score_simple_header(
            h, ('रकबा', 'area'), ('कुल', 'total', 'अर्जित', 'acquired', 'एकड़', 'acre'),
        ) + (5 if 'हे' in h or 'hectare' in h else 0)),
        'acquired_ha': _best_column(lambda h: _score_simple_header(
            h, ('अर्जित', 'acquired'), ('कुल', 'total', 'एकड़', 'acre'),
        ) + (5 if 'हे' in h or 'hectare' in h else 0)),
        'khata_total_ha': _best_column(lambda h: _score_simple_header(
            h, ('कुल', 'total'), ('एकड़', 'acre', 'अर्जित', 'acquired'),
        ) + (5 if 'हे' in h or 'hectare' in h else 0)),
        'khata_total_acre': _best_column(lambda h: _score_simple_header(
            h, ('कुल', 'total', 'एकड़', 'acre'), ('हे', 'hectare', 'अर्जित', 'acquired'),
        )),
        'acq_reference': _best_column(lambda h: _score_simple_header(
            h, ('भू-अर्जन', 'arjan', 'अर्जन वर्ष', 'acquisition'),
        )),
        'caste': _best_column(lambda h: _score_simple_header(h, ('जाति', 'caste', 'वर्ग'))),
        'employment': _best_column(lambda h: _score_simple_header(h, ('रोजगार', 'employment'))),
    }

    for key, default_idx in _DEFAULT_COL_MAP.items():
        if col_map.get(key) is None:
            col_map[key] = default_idx

    if col_map['khasra'] == col_map.get('owner'):
        for col_idx, header in enumerate(headers):
            if _score_khasra_header(header) >= 12:
                col_map['khasra'] = col_idx
                break

    if col_map['khasra'] is None:
        raise ValidationError(
            'Could not find the khasra/plot column (प्लांट नं./खसरा नं.) in the Excel header row.'
        )
    return col_map


def _looks_like_person_name(text):
    raw = _cell_text(text)
    if not raw:
        return False
    upper = raw.upper()
    if any(token in upper for token in (' D/O ', ' S/O ', ' W/O ')):
        return True
    if re.search(r'\s+(?:पिता|पति)\s+', raw):
        return True
    if re.search(r'\d', raw):
        return False
    return len(raw.split()) >= 3


def _looks_like_khasra_number(text):
    raw = _cell_text(text)
    if not raw or _looks_like_person_name(raw):
        return False
    return bool(re.match(r'^[\d./\-]+$', raw)) or bool(re.search(r'\d', raw))


def parse_landowner_text(text):
    """Split owner cell into name, father name, and spouse name."""
    raw = _cell_text(text)
    if not raw:
        return {'name': '', 'father_name': '', 'spouse_name': ''}

    split_patterns = (
        (r'\s+पिता\s+', 'father_name'),
        (r'\s+पति\s+', 'spouse_name'),
        (r'\s+(?:s/o|son of)\s+', 'father_name'),
        (r'\s+(?:d/o|daughter of)\s+', 'father_name'),
        (r'\s+(?:w/o|wife of)\s+', 'spouse_name'),
    )
    for pattern, relation_field in split_patterns:
        parts = re.split(pattern, raw, maxsplit=1, flags=re.IGNORECASE)
        if len(parts) == 2:
            result = {'name': parts[0].strip(), 'father_name': '', 'spouse_name': ''}
            result[relation_field] = parts[1].strip()
            return result
    return {'name': raw, 'father_name': '', 'spouse_name': ''}


def resolve_irrigation_type(text):
    raw = _cell_text(text).lower()
    if not raw:
        return 'unirrigated'
    if 'irrigated' in raw and 'unirrigated' not in raw and 'असिंचित' not in raw:
        return 'irrigated'
    if 'सिंचित' in raw or raw in ('irrigated', 'sinchit'):
        return 'irrigated'
    if 'असिंचित' in raw or raw in ('unirrigated', 'asinchit'):
        return 'unirrigated'
    return 'unirrigated'


def _irrigation_type_from_land_type(land_type):
    if not land_type:
        return 'unirrigated'
    name = (land_type.name or '').lower()
    code = (land_type.code or '').lower()
    if (
        'asin' in name or 'unirrig' in name or 'असिंचित' in (land_type.name or '')
        or 'asin' in code or code in ('002', 'asinchit', 'unirrigated')
    ):
        return 'unirrigated'
    if (
        'sinch' in name or 'irrig' in name or 'सिंचित' in (land_type.name or '')
        or 'sinch' in code or code in ('001', 'sinchit', 'irrigated')
    ):
        return 'irrigated'
    return resolve_irrigation_type(land_type.name or land_type.code or '')


def resolve_land_type_id(env, text):
    """Find or create a land type master record from Excel value."""
    raw = _cell_text(text)
    if not raw:
        return False, 'unirrigated'

    LandType = env['bhu.land.type'].sudo()
    record = LandType.search([
        '|', '|',
        ('name', '=ilike', raw),
        ('code', '=ilike', raw),
        ('name', 'ilike', raw),
    ], limit=1)

    if not record:
        normalized = raw.lower()
        if 'सिंचित' in raw or 'sinchit' in normalized or normalized == 'irrigated':
            record = LandType.search([
                '|', ('name', 'ilike', 'सिंचित'), ('name', 'ilike', 'sinchit'),
            ], limit=1)
        elif 'असिंचित' in raw or 'asinchit' in normalized or normalized == 'unirrigated':
            record = LandType.search([
                '|', ('name', 'ilike', 'असिंचित'), ('name', 'ilike', 'asinchit'),
            ], limit=1)

    if not record:
        slug = re.sub(r'[^a-zA-Z0-9]+', '_', raw).strip('_').upper()[:20] or 'LAND_TYPE'
        code = slug
        suffix = 1
        while LandType.search_count([('code', '=', code)]):
            suffix += 1
            code = f'{slug}_{suffix}'
        record = LandType.create({'name': raw, 'code': code})

    return record.id, _irrigation_type_from_land_type(record)


def _build_merged_value_map(sheet):
    merged = {}
    for merge_range in sheet.merged_cells.ranges:
        value = sheet.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged[(row, col)] = value
    return merged


def _sheet_cell(sheet, merged_map, row_idx, col_idx):
    key = (row_idx, col_idx + 1)
    if key in merged_map:
        return merged_map[key]
    return sheet.cell(row_idx, col_idx + 1).value


def _detect_header_row(sheet, max_scan=15):
    for row_idx in range(1, min(max_scan, sheet.max_row) + 1):
        texts = []
        for col_idx in range(sheet.max_column):
            texts.append(_cell_text(sheet.cell(row_idx, col_idx + 1).value).lower())
        joined = ' '.join(texts)
        if any(marker in joined for marker in _HEADER_MARKERS):
            return row_idx
    return 1


def parse_acquisition_xlsx(file_content, filename=''):
    """Return normalized row dicts from the acquisition schedule workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError(
            "Python library 'openpyxl' is required to import .xlsx files."
        ) from exc

    if not (filename or '').lower().endswith('.xlsx'):
        raise ValidationError('Please upload an .xlsx file.')

    workbook = load_workbook(io.BytesIO(file_content), data_only=True)
    sheet = workbook.active
    if not sheet or sheet.max_row < 2:
        raise ValidationError('The workbook is empty.')

    merged_map = _build_merged_value_map(sheet)
    header_row = _detect_header_row(sheet)
    col_map = _detect_column_map(sheet, header_row)
    data_start = header_row + 1

    rows = []
    master = {
        'khata_no': '',
        'owner_text': '',
        'owner_name': '',
        'father_name': '',
        'spouse_name': '',
        'caste': '',
        'khata_total_ha': 0.0,
        'khata_total_acre': 0.0,
        'employment_count': '',
        'acquisition_reference': '',
    }

    for row_idx in range(data_start, sheet.max_row + 1):
        khata_val = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['khata']))
        owner_val = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['owner']))
        caste_val = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['caste']))
        khata_total_ha = _cell_float(_sheet_cell(sheet, merged_map, row_idx, col_map['khata_total_ha']))
        khata_total_acre = _cell_float(_sheet_cell(sheet, merged_map, row_idx, col_map['khata_total_acre']))
        acq_ref = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['acq_reference']))
        employment = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['employment']))

        if khata_val:
            if master['khata_no'] and master['khata_no'] != khata_val and not owner_val:
                master.update({
                    'owner_text': '',
                    'owner_name': '',
                    'father_name': '',
                    'spouse_name': '',
                    'caste': '',
                })
            master['khata_no'] = khata_val
        if owner_val:
            master['owner_text'] = owner_val
            owner_parts = parse_landowner_text(owner_val)
            master['owner_name'] = owner_parts['name']
            master['father_name'] = owner_parts['father_name']
            master['spouse_name'] = owner_parts['spouse_name']
        if caste_val:
            master['caste'] = caste_val
        if khata_total_ha > 0:
            master['khata_total_ha'] = khata_total_ha
        if khata_total_acre > 0:
            master['khata_total_acre'] = khata_total_acre
        if acq_ref:
            master['acquisition_reference'] = acq_ref
        if employment:
            master['employment_count'] = employment

        khasra_number = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['khasra']))
        if not khasra_number:
            continue

        total_area = _cell_float(_sheet_cell(sheet, merged_map, row_idx, col_map['area_ha']))
        acquired_area = _cell_float(_sheet_cell(sheet, merged_map, row_idx, col_map['acquired_ha']))
        if acquired_area <= 0 and total_area > 0:
            acquired_area = total_area
        if total_area <= 0 and acquired_area > 0:
            total_area = acquired_area

        land_type = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['land_type']))
        row_acq_ref = _cell_text(_sheet_cell(sheet, merged_map, row_idx, col_map['acq_reference']))
        land_acquire_year = row_acq_ref or master['acquisition_reference']
        owner_name = master['owner_name'] or master['owner_text']
        if not owner_name:
            owner_name = 'Unknown'

        remarks_parts = []
        if master['employment_count']:
            remarks_parts.append(f"Employment: {master['employment_count']}")
        if master['khata_total_acre'] > 0:
            remarks_parts.append(f"Khata total (acre): {master['khata_total_acre']}")

        rows.append({
            'row_number': row_idx,
            'khata_no': master['khata_no'],
            'khasra_number': khasra_number,
            'total_area': total_area,
            'acquired_area': acquired_area,
            'land_type_text': land_type,
            'land_acquire_year': land_acquire_year,
            'landowner_name': owner_name,
            'father_name': master['father_name'],
            'spouse_name': master['spouse_name'],
            'caste': master['caste'],
            'khata_total_ha': master['khata_total_ha'],
            'remarks': ' | '.join(remarks_parts),
        })

    if not rows:
        raise ValidationError(
            'No survey rows found. Check that khasra/plot numbers exist in column "प्लांट नं./खसरा नं.".'
        )
    return rows


def import_survey_rows(env, rows, project_id, village_id, department_id,
                       area_id=False, update_existing=False, dry_run=False):
    """Create or update surveys from parsed acquisition rows."""
    Survey = env['bhu.survey'].sudo()
    project = env['bhu.project'].sudo().browse(project_id)
    village = env['bhu.village'].sudo().browse(village_id)
    department = env['bhu.department'].sudo().browse(department_id)
    area = env['bhukhadan.area.master'].sudo().browse(area_id) if area_id else env['bhukhadan.area.master']

    if not project.exists():
        raise ValidationError('Selected project does not exist.')
    if not village.exists():
        raise ValidationError('Selected village does not exist.')
    if not department.exists():
        raise ValidationError('Selected department does not exist.')
    if area_id and not area.exists():
        raise ValidationError('Selected area does not exist.')
    if project.village_ids and village not in project.village_ids:
        raise ValidationError(
            f'Village "{village.display_name}" is not linked to project "{project.name}".'
        )

    stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
    log_lines = []

    for row in rows:
        row_no = row['row_number']
        khasra = row['khasra_number']
        prefix = f'Row {row_no} (Khasra {khasra})'

        if _looks_like_person_name(khasra):
            stats['errors'] += 1
            log_lines.append(
                f'{prefix}: ERROR - "{khasra}" looks like landowner name; '
                'expected khasra from column "प्लांट नं./खसरा नं."'
            )
            continue
        if not _looks_like_khasra_number(khasra):
            stats['errors'] += 1
            log_lines.append(
                f'{prefix}: ERROR - invalid khasra "{khasra}"; expected values like 3, 22, or 240/1'
            )
            continue

        if row['total_area'] <= 0 or row['acquired_area'] <= 0:
            stats['errors'] += 1
            log_lines.append(f'{prefix}: ERROR - area must be greater than 0')
            continue
        if row['acquired_area'] > row['total_area']:
            stats['errors'] += 1
            log_lines.append(f'{prefix}: ERROR - acquired area exceeds total area')
            continue

        existing = Survey.search([
            ('village_id', '=', village.id),
            ('khasra_number', '=', khasra),
        ], limit=1)

        land_type_id, irrigation_type = resolve_land_type_id(env, row.get('land_type_text'))

        payload = {
            'project_id': project.id,
            'village_id': village.id,
            'department_id': department.id,
            'area_id': area.id if area_id else False,
            'khasra_number': khasra,
            'khata_no': row.get('khata_no') or False,
            'land_acquire_year': row.get('land_acquire_year') or False,
            'total_area': row['total_area'],
            'acquired_area': row['acquired_area'],
            'crop_type_id': land_type_id,
            'irrigation_type': irrigation_type,
            'survey_date': date.today().isoformat(),
            'remarks': row.get('remarks') or '',
            'landowners': [{
                'name': row['landowner_name'],
                'father_name': row.get('father_name') or False,
                'spouse_name': row.get('spouse_name') or False,
                'caste': row.get('caste') or False,
                'village_id': village.id,
            }],
        }

        try:
            if existing:
                if not update_existing:
                    stats['skipped'] += 1
                    log_lines.append(f'{prefix}: SKIPPED - survey already exists')
                    continue
                if existing.state not in ('draft', 'submitted'):
                    stats['errors'] += 1
                    log_lines.append(
                        f'{prefix}: ERROR - existing survey is {existing.state}; cannot update'
                    )
                    continue
                if dry_run:
                    stats['updated'] += 1
                    log_lines.append(f'{prefix}: OK (would update existing survey {existing.name})')
                    continue

                write_vals = {
                    'khata_no': payload.get('khata_no') or False,
                    'land_acquire_year': payload.get('land_acquire_year') or False,
                    'area_id': area.id if area_id else False,
                    'total_area': payload['total_area'],
                    'acquired_area': payload['acquired_area'],
                    'crop_type_id': land_type_id or False,
                    'irrigation_type': irrigation_type,
                    'remarks': payload['remarks'],
                    'project_id': project.id,
                    'department_id': department.id,
                }
                existing.write(write_vals)
                if existing.landowner_ids:
                    existing.landowner_ids[0].write({
                        'name': row['landowner_name'],
                        'father_name': row.get('father_name') or False,
                        'spouse_name': row.get('spouse_name') or False,
                        'caste': row.get('caste') or False,
                        'village_id': village.id,
                    })
                else:
                    existing.write({
                        'landowner_ids': [(0, 0, payload['landowners'][0])],
                    })
                stats['updated'] += 1
                log_lines.append(f'{prefix}: UPDATED survey {existing.name}')
                continue

            if dry_run:
                stats['created'] += 1
                log_lines.append(f'{prefix}: OK (would create new survey)')
                continue

            vals = api_build_survey_vals(env, payload, user_id=env.user.id)
            survey = Survey.create(vals)
            stats['created'] += 1
            log_lines.append(f'{prefix}: CREATED survey {survey.name}')
        except ValidationError as err:
            stats['errors'] += 1
            log_lines.append(f'{prefix}: ERROR - {err.args[0]}')
        except Exception as err:
            stats['errors'] += 1
            log_lines.append(f'{prefix}: ERROR - {err}')

    summary = (
        f"Created: {stats['created']} | Updated: {stats['updated']} | "
        f"Skipped: {stats['skipped']} | Errors: {stats['errors']}"
    )
    log_lines.insert(0, summary)
    return stats, '\n'.join(log_lines)


def import_survey_xlsx(env, file_content, filename, project_id, village_id, department_id,
                       area_id=False, update_existing=False, dry_run=False):
    rows = parse_acquisition_xlsx(base64.b64decode(file_content), filename=filename)
    return import_survey_rows(
        env, rows, project_id, village_id, department_id,
        area_id=area_id, update_existing=update_existing, dry_run=dry_run,
    )
